# -*- coding: utf-8 -*-
import openpyxl
import os

excel_path = 'K572+774红石牡丹江大桥（右幅）病害.xls'
if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    # 检查第一张表
    ws = wb.active
    print('工作表名称:', wb.sheetnames)
    # 检查前几行
    for row in range(1, 10):
        for col in range(1, 5):
            cell = ws.cell(row, col)
            if cell.value and 'K572' in str(cell.value):
                print(f'({row},{col}): {repr(cell.value)}')
                print(f'  类型: {type(cell.value)}')
    wb.close()
